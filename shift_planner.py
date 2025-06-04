import random
import csv
from openpyxl import Workbook, load_workbook

class Staff: 
    def __init__(self, staff_id: int, total_days: int):
        self.id = staff_id;
        self.schedule = [False] * total_days # False = day off
        
    def assign_shift(self, day: int):
        self.schedule[day] = True 
        
    def days_off(self) -> int:
        return self.schedule.count(False)
    
    def weekend_shifts(self ) -> int:
        # Count how many weekend days (5 = saturday, 6 = sunday) are work days
        return sum(1 for i in range(len(self.schedule)) if self.schedule[i] and i % 7 in [5, 6])
    
class ShiftPlanner: 
    def __init__(self, m: int, n: int, k: int, daily_requirements: list[int], history: list[list[list[str]]]):
        self.m = m
        self.n = n
        self.k = k
        self.daily_requirements = daily_requirements
        self.staff_members = [Staff(i, n) for i in range(m)]
        self.schedule_matrix = [[False for _ in range(n)] for _ in range(m)]
        self.history = history # List of previous weeks' schedules matrixes
    
    def get_weekend_load(self, staff_id: int) -> int:
        count = 0
        for week in self.history:
            for d in [5, 6]:
                if d < len(week) and week[staff_id][d].upper() == "WORK":
                    count += 1
        return count
        
    def generate_schedule(self):
        for day in range(self.n):
            min_required = self.daily_requirements[day]
            target = min(self.m, min_required + 1) 
            available_staff = list(range(self.m))
            if day % 7 in [5, 6]:
                # Weekend day, prioritize staff with fewer weekend shifts
                available_staff.sort(
                                      key=lambda sid: self.get_weekend_load(sid) +
                                                       self.staff_members[sid].weekend_shifts())
            else:
                random.shuffle(available_staff) 
            
            assigned_today = 0
            for sid in available_staff:
                staff = self.staff_members[sid]
                if staff.days_off() > self.k:
                    staff.assign_shift(day)
                    self.schedule_matrix[sid][day] = True
                    assigned_today += 1
                    if assigned_today >= target:
                        break;
                        
        for staff in self.staff_members:
            working_days = staff.schedule.count(True)
            if self.n - working_days < self.k:
                #Remove random work days to ensure k days off
                to_remove = self.k - (self.n - working_days)
                work_days = [i for i, worked in enumerate(staff.schedule) if worked]
                random.shuffle(work_days)
                for i in range(to_remove):
                    day = work_days[i]
                    staff.schedule[day] = False
                    self.schedule_matrix[staff.id][day] = False
                    
    def handle_leave_request(self, staff_id: int, day: int):
            staff = self.staff_members[staff_id]
            if not staff.schedule[day]:
                print(f"Staff {staff_id} is already off on Day {day+1}")
                return
            
            # Mark as day off
            staff.schedule[day] = False
            self.schedule_matrix[staff_id][day] = False
            print(f"Marked Staff {staff_id}'s Day { day+1} as off (leave request).")
            
            #Reassign another staff member for that day if needed
            current_staff = [s for s in self.staff_members if s.schedule[day]]
            if len(current_staff) < self.daily_requirements[day]:
                #Find replacement if necessary 
                for s in self.staff_members:
                    if not s.schedule[day] and s.days_off() > self.k:
                        s.schedule[day] = True
                        self.schedule_matrix[s.id][day] = True
                        print(f"Assigned staff {s.id} to cover Day {day+1}")
                        return
                print("Warning: Could not find replacement to meet daily requirement.")
        
    def display_schedule(self):
        print("\n=== Schedule Matrix ===")
        
        header = ["Staff ID"] + [f"Day { i+1}" for i in range(self.n)]
        print("{:<10}".format(header[0]), end="")
        for day in header[1:]:
            print("{:<8}".format(day), end="")
        print()
        
        for staff in self.staff_members:
            print("{:<10}".format(f"Staff { staff.id:02d}"), end="")
            for day in staff.schedule:
                print("{:<8}".format("WORK" if day else "OFF"), end="")
            print()
            
    def validate_schedule(self):
        valid_staff = all(staff.days_off() >= self.k for staff in self.staff_members)
        valid_days = True
        for day in range(self.n):
            working = sum(staff.schedule[day] for staff in self.staff_members)
            if working < self.daily_requirements[day]: 
                print(f"Day {day + 1} failed: has {working}, needs {self.daily_requirements[day]}")
                valid_days = False
        return valid_staff and valid_days
        
    def export_schedule_csv(self, filename = "schedule.csv"):
            with open(filename, mode = 'w', newline = '') as file:
                writer = csv.writer(file)
                header = ["Staff ID"] + [f"Day {i+1}" for i in range(self.n)]
                writer.writerow(header)
                for staff in self.staff_members: 
                    row = [f"Staff {staff.id}"] + ["WORK" if x else "OFF" for x in staff.schedule]
                    writer.writerow(row)
                print(f"Schedule exported to {filename}")
    
    def export_schedule_excel(self, filename = "schedule.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(["Staff ID"] + [f"Day {i+1}" for i in range(self.n)])
        for staff in self.staff_members:
            row = [f"Staff {staff.id}"] + ["WORK" if x else "OFF" for x in staff.schedule]
            ws.append(row)
        wb.save(filename)
        print(f"Schedule exported to {filename}")

def load_previous_weeks(file_paths: list[str], m, n) -> list[list[list[str]]]:
    history = []
    for path in file_paths:
        week_schedule = []
        try:
            if path.endswith('.csv'):
                with open(path, 'r') as file:
                    reader = csv.reader(file)
                    for row in reader:
                        week_schedule.append(row)
            elif path.endswith('.xlsx'):
                wb = load_workbook(path)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    week_schedule.append([str(cell) if cell is not None else "" for cell in row])
            else:
                print(f"Unsupported file type: {path}")
                continue
            history.append(week_schedule)
        except Exception as e:
            print(f"Error loading {path}: {e}")
    return history